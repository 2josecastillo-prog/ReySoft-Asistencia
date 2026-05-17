from collections.abc import Iterable
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import Organization
from app.schemas.report import AttendanceCourseReport, AttendanceReportRecord, AttendanceStudentReport

REPORT_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME_TYPE = "application/pdf"

STATUS_LABELS = {
    "arrived": "Llegó",
    "absent": "Ausente",
    "late": "Tarde",
    "early_pickup": "Retiro temprano",
    "excused": "Excusado",
}

RISK_LABELS = {
    "ok": "Bajo",
    "warning": "Atención",
    "danger": "Alto",
}

RISK_FILLS = {
    "ok": "FFD1FAE5",
    "warning": "FFFEF3C7",
    "danger": "FFFEE2E2",
}

RISK_FONTS = {
    "ok": "FF047857",
    "warning": "FF92400E",
    "danger": "FFB91C1C",
}

THIN_BORDER = Border(bottom=Side(style="thin", color="FFE2E8F0"))


def _argb(hex_color: str | None, fallback: str = "2563EB") -> str:
    cleaned = (hex_color or fallback).strip().lstrip("#")
    if len(cleaned) != 6:
        cleaned = fallback
    return f"FF{cleaned.upper()}"


def _pdf_color(hex_color: str | None, fallback: str = "2563EB"):
    cleaned = (hex_color or fallback).strip()
    if not cleaned.startswith("#"):
        cleaned = f"#{cleaned}"
    if len(cleaned) != 7:
        cleaned = f"#{fallback}"
    return HexColor(cleaned)


def _course_label(name: str, section: str | None, academic_year: str | None) -> str:
    return " ".join(part for part in (name, section, academic_year) if part)


def _period_label(start_date: date | None, end_date: date | None) -> str:
    if start_date and end_date:
        return f"Periodo: {start_date.isoformat()} al {end_date.isoformat()}"
    if start_date:
        return f"Periodo: desde {start_date.isoformat()}"
    if end_date:
        return f"Periodo: hasta {end_date.isoformat()}"
    return "Periodo: todos los registros"


def _time_label(value: time | None) -> str | None:
    return value.strftime("%H:%M") if value else None


def _prepare_sheet(workbook: Workbook, title: str):
    worksheet = workbook.create_sheet(title)
    worksheet.sheet_view.showGridLines = False
    return worksheet


def _write_report_header(worksheet, organization: Organization, start_date: date | None, end_date: date | None, columns: int) -> None:
    last_column = get_column_letter(columns)
    title_range = f"A1:{last_column}1"
    organization_range = f"A2:{last_column}2"
    meta_range = f"A3:{last_column}3"
    generated_range = f"A4:{last_column}4"

    worksheet.merge_cells(title_range)
    worksheet.merge_cells(organization_range)
    worksheet.merge_cells(meta_range)
    worksheet.merge_cells(generated_range)
    worksheet["A1"] = "Reporte institucional de asistencia"
    worksheet["A2"] = organization.name
    worksheet["A3"] = _period_label(start_date, end_date)
    worksheet["A4"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    primary = _argb(organization.primary_color)
    secondary = _argb(organization.secondary_color, fallback="1E293B")
    worksheet["A1"].font = Font(bold=True, size=16, color=primary)
    worksheet["A2"].font = Font(bold=True, size=12, color=secondary)
    worksheet["A3"].font = Font(size=10, color="FF475569")
    worksheet["A4"].font = Font(size=10, color="FF475569")


def _write_table_header(worksheet, row_index: int, headers: list[str], organization: Organization) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=_argb(organization.primary_color))
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_table_rows(worksheet, min_row: int, max_column: int) -> None:
    for row in worksheet.iter_rows(min_row=min_row, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _style_risk_cell(cell, risk_level: str) -> None:
    cell.value = RISK_LABELS.get(risk_level, risk_level)
    cell.fill = PatternFill(fill_type="solid", fgColor=RISK_FILLS.get(risk_level, "FFFFFFFF"))
    cell.font = Font(bold=True, color=RISK_FONTS.get(risk_level, "FF0F172A"))
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(worksheet, max_column_width: int = 48) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), max_column_width)


def _write_detail_sheet(
    worksheet,
    organization: Organization,
    records: Iterable[tuple[AttendanceReportRecord, str]],
    start_date: date | None,
    end_date: date | None,
) -> None:
    headers = ["Fecha", "Estudiante", "Curso", "Estado", "Hora", "Notas"]
    _write_report_header(worksheet, organization, start_date, end_date, len(headers))
    _write_table_header(worksheet, 6, headers, organization)
    for record, course_label in records:
        worksheet.append(
            [
                record.attendance_date.isoformat(),
                record.student_name,
                course_label,
                STATUS_LABELS.get(record.status, record.status),
                _time_label(record.display_time),
                record.notes,
            ]
        )
    worksheet.freeze_panes = "A7"
    worksheet.auto_filter.ref = f"A6:F{worksheet.max_row}"
    _style_table_rows(worksheet, 7, len(headers))
    _autosize_columns(worksheet)


def _workbook_bytes(workbook: Workbook) -> BytesIO:
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _pdf_styles(organization: Organization):
    base = getSampleStyleSheet()
    primary = _pdf_color(organization.primary_color)
    secondary = _pdf_color(organization.secondary_color, fallback="1E293B")
    return {
        "title": ParagraphStyle(
            "ReportTitle",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=primary,
            alignment=TA_LEFT,
            spaceAfter=4,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=secondary,
            alignment=TA_LEFT,
        ),
        "meta": ParagraphStyle(
            "ReportMeta",
            parent=base["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569"),
        ),
        "cell": ParagraphStyle(
            "ReportCell",
            parent=base["Normal"],
            fontSize=7,
            leading=9,
            alignment=TA_LEFT,
        ),
        "center_cell": ParagraphStyle(
            "ReportCenterCell",
            parent=base["Normal"],
            fontSize=7,
            leading=9,
            alignment=TA_CENTER,
        ),
    }


def _paragraph(value, style: ParagraphStyle) -> Paragraph:
    text = "" if value is None else str(value)
    return Paragraph(text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), style)


def _pdf_table(data: list[list], organization: Organization, widths: list[float]) -> Table:
    table = Table(data, repeatRows=1, colWidths=widths)
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_color(organization.primary_color)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]
    )
    table.setStyle(style)
    return table


def _add_pdf_header(story: list, organization: Organization, start_date: date | None, end_date: date | None) -> None:
    styles = _pdf_styles(organization)
    story.append(_paragraph("Reporte institucional de asistencia", styles["title"]))
    story.append(_paragraph(organization.name, styles["subtitle"]))
    story.append(_paragraph(_period_label(start_date, end_date), styles["meta"]))
    story.append(_paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["meta"]))
    story.append(Spacer(1, 0.28 * cm))


def _risk_label_paragraph(risk_level: str, styles: dict[str, ParagraphStyle]) -> Paragraph:
    return _paragraph(RISK_LABELS.get(risk_level, risk_level), styles["center_cell"])


def _build_pdf(story: list) -> BytesIO:
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )
    document.build(story)
    stream.seek(0)
    return stream


def build_student_report_pdf(
    organization: Organization,
    rows: list[AttendanceStudentReport],
    start_date: date | None,
    end_date: date | None,
) -> BytesIO:
    styles = _pdf_styles(organization)
    story: list = []
    _add_pdf_header(story, organization, start_date, end_date)

    summary_headers = ["Estudiante", "Codigo", "Curso", "Aus. eq.", "Aus.", "Exc.", "Conv.", "Tardes", "Nivel", "Reg."]
    summary_data = [[_paragraph(header, styles["center_cell"]) for header in summary_headers]]
    detail_data = [[_paragraph(header, styles["center_cell"]) for header in ["Fecha", "Estudiante", "Curso", "Estado", "Hora", "Notas"]]]
    for row in rows:
        course = _course_label(row.course_name, row.course_section, row.course_academic_year)
        summary_data.append(
            [
                _paragraph(row.student_name, styles["cell"]),
                _paragraph(row.student_code, styles["center_cell"]),
                _paragraph(course, styles["cell"]),
                _paragraph(row.equivalent_absences, styles["center_cell"]),
                _paragraph(row.absent_count, styles["center_cell"]),
                _paragraph(row.excused_count, styles["center_cell"]),
                _paragraph(row.excused_absence_equivalent, styles["center_cell"]),
                _paragraph(row.late_count, styles["center_cell"]),
                _risk_label_paragraph(row.risk_level, styles),
                _paragraph(row.total_records, styles["center_cell"]),
            ]
        )
        for record in row.records:
            detail_data.append(
                [
                    _paragraph(record.attendance_date.isoformat(), styles["center_cell"]),
                    _paragraph(record.student_name, styles["cell"]),
                    _paragraph(course, styles["cell"]),
                    _paragraph(STATUS_LABELS.get(record.status, record.status), styles["center_cell"]),
                    _paragraph(_time_label(record.display_time), styles["center_cell"]),
                    _paragraph(record.notes, styles["cell"]),
                ]
            )

    story.append(_pdf_table(summary_data, organization, [4.5 * cm, 2.0 * cm, 4.0 * cm, 1.6 * cm, 1.3 * cm, 1.3 * cm, 1.4 * cm, 1.5 * cm, 1.7 * cm, 1.2 * cm]))
    story.append(Spacer(1, 0.35 * cm))
    story.append(_paragraph("Detalle de asistencia", styles["subtitle"]))
    story.append(_pdf_table(detail_data, organization, [2.2 * cm, 4.5 * cm, 4.2 * cm, 2.5 * cm, 1.7 * cm, 9.0 * cm]))
    return _build_pdf(story)


def build_course_report_pdf(
    organization: Organization,
    rows: list[AttendanceCourseReport],
    start_date: date | None,
    end_date: date | None,
) -> BytesIO:
    styles = _pdf_styles(organization)
    story: list = []
    _add_pdf_header(story, organization, start_date, end_date)

    summary_headers = ["Curso", "Est.", "Aus. eq.", "Aus.", "Exc.", "Conv.", "Tardes", "Atencion", "Alto", "Nivel", "Reg."]
    summary_data = [[_paragraph(header, styles["center_cell"]) for header in summary_headers]]
    detail_data = [[_paragraph(header, styles["center_cell"]) for header in ["Fecha", "Estudiante", "Curso", "Estado", "Hora", "Notas"]]]
    for row in rows:
        course = _course_label(row.course_name, row.course_section, row.course_academic_year)
        summary_data.append(
            [
                _paragraph(course, styles["cell"]),
                _paragraph(row.student_count, styles["center_cell"]),
                _paragraph(row.equivalent_absences, styles["center_cell"]),
                _paragraph(row.absent_count, styles["center_cell"]),
                _paragraph(row.excused_count, styles["center_cell"]),
                _paragraph(row.excused_absence_equivalent, styles["center_cell"]),
                _paragraph(row.late_count, styles["center_cell"]),
                _paragraph(row.warning_students, styles["center_cell"]),
                _paragraph(row.danger_students, styles["center_cell"]),
                _risk_label_paragraph(row.risk_level, styles),
                _paragraph(row.total_records, styles["center_cell"]),
            ]
        )
        for record in row.records:
            detail_data.append(
                [
                    _paragraph(record.attendance_date.isoformat(), styles["center_cell"]),
                    _paragraph(record.student_name, styles["cell"]),
                    _paragraph(course, styles["cell"]),
                    _paragraph(STATUS_LABELS.get(record.status, record.status), styles["center_cell"]),
                    _paragraph(_time_label(record.display_time), styles["center_cell"]),
                    _paragraph(record.notes, styles["cell"]),
                ]
            )

    story.append(_pdf_table(summary_data, organization, [4.8 * cm, 1.2 * cm, 1.6 * cm, 1.2 * cm, 1.2 * cm, 1.4 * cm, 1.4 * cm, 1.7 * cm, 1.2 * cm, 1.7 * cm, 1.2 * cm]))
    story.append(Spacer(1, 0.35 * cm))
    story.append(_paragraph("Detalle de asistencia", styles["subtitle"]))
    story.append(_pdf_table(detail_data, organization, [2.2 * cm, 4.5 * cm, 4.2 * cm, 2.5 * cm, 1.7 * cm, 9.0 * cm]))
    return _build_pdf(story)


def build_student_report_workbook(
    organization: Organization,
    rows: list[AttendanceStudentReport],
    start_date: date | None,
    end_date: date | None,
) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen por estudiante"
    summary.sheet_view.showGridLines = False

    headers = [
        "Estudiante",
        "Código",
        "Curso",
        "Ausencias eq.",
        "Ausencias",
        "Excusas",
        "Conv. excusas",
        "Tardes",
        "Nivel",
        "Registros",
    ]
    _write_report_header(summary, organization, start_date, end_date, len(headers))
    _write_table_header(summary, 6, headers, organization)

    detail_records: list[tuple[AttendanceReportRecord, str]] = []
    for row in rows:
        course = _course_label(row.course_name, row.course_section, row.course_academic_year)
        summary.append(
            [
                row.student_name,
                row.student_code,
                course,
                row.equivalent_absences,
                row.absent_count,
                row.excused_count,
                row.excused_absence_equivalent,
                row.late_count,
                RISK_LABELS.get(row.risk_level, row.risk_level),
                row.total_records,
            ]
        )
        _style_risk_cell(summary.cell(row=summary.max_row, column=9), row.risk_level)
        for record in row.records:
            detail_records.append((record, course))

    summary.freeze_panes = "A7"
    summary.auto_filter.ref = f"A6:J{summary.max_row}"
    _style_table_rows(summary, 7, len(headers))
    _autosize_columns(summary)

    detail = _prepare_sheet(workbook, "Detalle")
    _write_detail_sheet(detail, organization, detail_records, start_date, end_date)
    return _workbook_bytes(workbook)


def build_course_report_workbook(
    organization: Organization,
    rows: list[AttendanceCourseReport],
    start_date: date | None,
    end_date: date | None,
) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen por curso"
    summary.sheet_view.showGridLines = False

    headers = [
        "Curso",
        "Estudiantes",
        "Ausencias eq.",
        "Ausencias",
        "Excusas",
        "Conv. excusas",
        "Tardes",
        "Atención",
        "Riesgo alto",
        "Nivel",
        "Registros",
    ]
    _write_report_header(summary, organization, start_date, end_date, len(headers))
    _write_table_header(summary, 6, headers, organization)

    detail_records: list[tuple[AttendanceReportRecord, str]] = []
    for row in rows:
        course = _course_label(row.course_name, row.course_section, row.course_academic_year)
        summary.append(
            [
                course,
                row.student_count,
                row.equivalent_absences,
                row.absent_count,
                row.excused_count,
                row.excused_absence_equivalent,
                row.late_count,
                row.warning_students,
                row.danger_students,
                RISK_LABELS.get(row.risk_level, row.risk_level),
                row.total_records,
            ]
        )
        _style_risk_cell(summary.cell(row=summary.max_row, column=10), row.risk_level)
        for record in row.records:
            detail_records.append((record, course))

    summary.freeze_panes = "A7"
    summary.auto_filter.ref = f"A6:K{summary.max_row}"
    _style_table_rows(summary, 7, len(headers))
    _autosize_columns(summary)

    detail = _prepare_sheet(workbook, "Detalle")
    _write_detail_sheet(detail, organization, detail_records, start_date, end_date)
    return _workbook_bytes(workbook)

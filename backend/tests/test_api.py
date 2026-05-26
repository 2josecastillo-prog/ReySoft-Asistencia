import csv
import json
from datetime import date, timedelta
from io import BytesIO, StringIO
from pathlib import Path

from fastapi.testclient import TestClient
from jose import jwt
from openpyxl import Workbook, load_workbook

from app.core.config import settings


def auth_headers(client: TestClient, email: str, password: str) -> dict[str, str]:
    response = client.post("/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_api_security_headers_and_sensitive_cache_policy(client: TestClient):
    response = client.get("/auth/me")

    assert response.status_code == 401
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-frame-options"] == "DENY"
    assert response.headers["referrer-policy"] == "strict-origin-when-cross-origin"
    assert "camera=()" in response.headers["permissions-policy"]
    assert "frame-ancestors 'none'" in response.headers["content-security-policy"]
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["pragma"] == "no-cache"

    mounted_response = client.get("/api/auth/me")
    assert mounted_response.headers["cache-control"] == "no-store"


def test_vercel_frontend_security_headers_are_configured():
    config_path = Path(__file__).resolve().parents[2] / "vercel.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    header_groups = config.get("headers", [])
    headers = {
        header["key"].lower(): header["value"]
        for group in header_groups
        for header in group.get("headers", [])
    }

    assert headers["x-content-type-options"] == "nosniff"
    assert headers["x-frame-options"] == "DENY"
    assert headers["referrer-policy"] == "strict-origin-when-cross-origin"
    assert "camera=()" in headers["permissions-policy"]
    assert "frame-ancestors 'none'" in headers["content-security-policy"]


def create_school_by_super_admin(client: TestClient, suffix: str = "one", status: str = "active") -> dict:
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    payload = {
        "organization_name": f"Colegio {suffix}",
        "organization_email": f"contacto-{suffix}@example.com",
        "organization_phone": "(809) 555-1234",
        "admin_first_name": "Admin",
        "admin_last_name": suffix.title(),
        "admin_email": f"admin-{suffix}@example.com",
        "password": "SchoolAdmin123!",
        "status": status,
        "primary_color": "#2563EB",
        "secondary_color": "#1E293B",
        "accent_color": "#F59E0B",
    }
    response = client.post("/admin/organizations", json=payload, headers=headers)
    assert response.status_code == 201, response.text
    return response.json()["organization"]


def activate_school(client: TestClient, organization_id: str) -> None:
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    response = client.post(
        f"/admin/organizations/{organization_id}/activate",
        json={"notes": "Pago externo confirmado"},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["status"] == "active"


def test_name_parts_are_stored_and_full_name_is_computed(client: TestClient):
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    response = client.post(
        "/admin/organizations",
        json={
            "organization_name": "Colegio Nombres",
            "organization_email": "nombres@example.com",
            "organization_phone": "(809) 555-1234",
            "admin_first_name": "Ana",
            "admin_middle_name": "Isabel",
            "admin_last_name": "Matos",
            "admin_second_surname": "Reyes",
            "admin_email": "admin-nombres@example.com",
            "password": "SchoolAdmin123!",
            "status": "active",
        },
        headers=headers,
    )
    assert response.status_code == 201, response.text
    admin_user = response.json()["admin_user"]
    assert admin_user["first_name"] == "Ana"
    assert admin_user["middle_name"] == "Isabel"
    assert admin_user["last_name"] == "Matos"
    assert admin_user["second_surname"] == "Reyes"
    assert admin_user["full_name"] == "Ana Isabel Matos Reyes"


def create_school_basics(client: TestClient, headers: dict[str, str]) -> tuple[dict, dict, dict]:
    course_response = client.post(
        "/courses",
        json={"name": "Primero", "section": "A", "academic_year": "2026-2027"},
        headers=headers,
    )
    assert course_response.status_code == 201, course_response.text
    guardian_response = client.post(
        "/guardians",
        json={"first_name": "María", "last_name": "Rodríguez", "phone": "(809) 555-1234", "relationship": "Madre"},
        headers=headers,
    )
    assert guardian_response.status_code == 201, guardian_response.text
    student_response = client.post(
        "/students",
        json={
            "first_name": "Luis",
            "last_name": "Pérez",
            "student_code": "ST-001",
            "course_id": course_response.json()["id"],
            "guardian_ids": [guardian_response.json()["id"]],
            "primary_guardian_id": guardian_response.json()["id"],
        },
        headers=headers,
    )
    assert student_response.status_code == 201, student_response.text
    return course_response.json(), guardian_response.json(), student_response.json()


def test_public_registration_is_disabled_and_super_admin_can_create_pending_school(client: TestClient):
    public_response = client.post(
        "/auth/register",
        json={
            "organization_name": "Colegio Publico",
            "organization_email": "publico@example.com",
            "organization_phone": "(809) 555-1234",
            "admin_first_name": "Admin",
            "admin_last_name": "Publico",
            "admin_email": "admin-publico@example.com",
            "password": "SchoolAdmin123!",
        },
    )
    assert public_response.status_code == 404

    organization = create_school_by_super_admin(client, status="pending")
    assert organization["status"] == "pending"

    response = client.post("/auth/login", json={"email": "admin-one@example.com", "password": "SchoolAdmin123!"})
    assert response.status_code == 403
    assert "pendiente de activación" in response.json()["detail"]


def test_super_admin_activates_school_and_school_admin_can_login(client: TestClient):
    organization = create_school_by_super_admin(client, status="pending")
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    response = client.get("/auth/me", headers=headers)
    assert response.status_code == 200
    assert response.json()["organization"]["status"] == "active"


def test_user_login_token_has_standard_security_claims(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])

    response = client.post("/auth/login", json={"email": "admin-one@example.com", "password": "SchoolAdmin123!"})
    assert response.status_code == 200, response.text

    claims = jwt.get_unverified_claims(response.json()["access_token"])
    assert claims["iss"] == settings.jwt_issuer
    assert claims["aud"] == settings.jwt_audience
    assert claims["typ"] == "access"
    assert claims["role"] == "school_admin"
    assert claims["token_version"] == 0
    assert isinstance(claims["jti"], str)
    assert isinstance(claims["iat"], int)
    assert isinstance(claims["nbf"], int)
    assert claims["exp"] - claims["iat"] == settings.access_token_expire_minutes * 60


def test_user_login_sets_http_only_cookie_and_logout_clears_it(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])

    login = client.post("/auth/login", json={"email": "admin-one@example.com", "password": "SchoolAdmin123!"})
    assert login.status_code == 200, login.text
    assert "reysoft_asistencia_token" in login.headers["set-cookie"]
    assert "reysoft_asistencia_csrf_token" in login.headers["set-cookie"]
    assert "HttpOnly" in login.headers["set-cookie"]
    csrf_token = client.cookies.get("reysoft_asistencia_csrf_token")
    assert csrf_token

    me = client.get("/auth/me")
    assert me.status_code == 200, me.text
    assert me.json()["email"] == "admin-one@example.com"

    logout = client.post("/auth/logout", headers={"X-CSRF-Token": csrf_token})
    assert logout.status_code == 200, logout.text
    assert "reysoft_asistencia_token" in logout.headers["set-cookie"]
    assert "reysoft_asistencia_csrf_token" in logout.headers["set-cookie"]
    assert "Max-Age=0" in logout.headers["set-cookie"]

    blocked = client.get("/auth/me")
    assert blocked.status_code == 401


def test_expired_activation_auto_suspends_school_and_blocks_school_login(client: TestClient):
    organization = create_school_by_super_admin(client, status="pending")
    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    expired_date = (date.today() - timedelta(days=1)).isoformat()
    activation = client.post(
        f"/admin/organizations/{organization['id']}/activate",
        json={"expiration_date": expired_date, "notes": "Pago vencido."},
        headers=super_admin_headers,
    )
    assert activation.status_code == 200, activation.text
    assert activation.json()["status"] == "active"

    blocked_login = client.post("/auth/login", json={"email": "admin-one@example.com", "password": "SchoolAdmin123!"})
    assert blocked_login.status_code == 403
    assert "expir" in blocked_login.json()["detail"].lower()

    organization_detail = client.get(f"/admin/organizations/{organization['id']}", headers=super_admin_headers)
    assert organization_detail.status_code == 200, organization_detail.text
    assert organization_detail.json()["status"] == "suspended"


def test_super_admin_marks_all_notifications_as_read(client: TestClient):
    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    first_organization = create_school_by_super_admin(client, "notify-one", status="pending")
    second_organization = create_school_by_super_admin(client, "notify-two", status="pending")

    for organization in [first_organization, second_organization]:
        response = client.post(
            f"/admin/organizations/{organization['id']}/activate",
            json={"notes": "Pago externo confirmado"},
            headers=super_admin_headers,
        )
        assert response.status_code == 200, response.text

    unread_notifications = client.get(
        "/admin/notifications",
        params={"unread_only": True},
        headers=super_admin_headers,
    )
    assert unread_notifications.status_code == 200, unread_notifications.text
    assert len(unread_notifications.json()) == 2

    mark_all = client.put("/admin/notifications/read-all", headers=super_admin_headers)
    assert mark_all.status_code == 200, mark_all.text
    assert mark_all.json() == {"updated_count": 2}

    remaining_unread = client.get(
        "/admin/notifications",
        params={"unread_only": True},
        headers=super_admin_headers,
    )
    assert remaining_unread.status_code == 200, remaining_unread.text
    assert remaining_unread.json() == []


def test_school_users_have_per_user_realtime_notifications(client: TestClient):
    first_organization = create_school_by_super_admin(client, "notify-school", status="pending")
    second_organization = create_school_by_super_admin(client, "notify-other", status="pending")
    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")

    for organization in [first_organization, second_organization]:
        response = client.post(
            f"/admin/organizations/{organization['id']}/activate",
            json={"notes": "Pago externo confirmado"},
            headers=super_admin_headers,
        )
        assert response.status_code == 200, response.text

    school_admin_headers = auth_headers(client, "admin-notify-school@example.com", "SchoolAdmin123!")
    staff_response = client.post(
        "/users",
        json={
            "first_name": "Personal",
            "last_name": "Notificaciones",
            "email": "staff-notify-school@example.com",
            "password": "Staff12345!",
        },
        headers=school_admin_headers,
    )
    assert staff_response.status_code == 201, staff_response.text
    staff_headers = auth_headers(client, "staff-notify-school@example.com", "Staff12345!")
    other_school_headers = auth_headers(client, "admin-notify-other@example.com", "SchoolAdmin123!")

    school_notifications = client.get("/notifications", headers=school_admin_headers)
    assert school_notifications.status_code == 200, school_notifications.text
    assert [item["organization_id"] for item in school_notifications.json()] == [first_organization["id"]]
    assert school_notifications.json()[0]["is_read"] is False

    staff_notifications = client.get("/notifications", headers=staff_headers)
    assert staff_notifications.status_code == 200, staff_notifications.text
    assert staff_notifications.json()[0]["id"] == school_notifications.json()[0]["id"]
    assert staff_notifications.json()[0]["is_read"] is False

    read_by_admin = client.put(
        f"/notifications/{school_notifications.json()[0]['id']}/read",
        headers=school_admin_headers,
    )
    assert read_by_admin.status_code == 200, read_by_admin.text
    assert read_by_admin.json()["is_read"] is True

    admin_unread = client.get("/notifications", params={"unread_only": True}, headers=school_admin_headers)
    assert admin_unread.status_code == 200, admin_unread.text
    assert admin_unread.json() == []

    staff_unread = client.get("/notifications", params={"unread_only": True}, headers=staff_headers)
    assert staff_unread.status_code == 200, staff_unread.text
    assert [item["id"] for item in staff_unread.json()] == [school_notifications.json()[0]["id"]]

    read_all_by_staff = client.put("/notifications/read-all", headers=staff_headers)
    assert read_all_by_staff.status_code == 200, read_all_by_staff.text
    assert read_all_by_staff.json() == {"updated_count": 1}

    other_school_notifications = client.get("/notifications", headers=other_school_headers)
    assert other_school_notifications.status_code == 200, other_school_notifications.text
    assert [item["organization_id"] for item in other_school_notifications.json()] == [second_organization["id"]]


def test_authenticated_user_can_open_notifications_websocket(client: TestClient):
    organization = create_school_by_super_admin(client, "notify-ws", status="active")
    activate_school(client, organization["id"])
    login = client.post("/auth/login", json={"email": "admin-notify-ws@example.com", "password": "SchoolAdmin123!"})
    assert login.status_code == 200, login.text

    with client.websocket_connect(f"/notifications/ws?token={login.json()['access_token']}") as websocket:
        websocket.send_text("ping")


def test_super_admin_lists_audit_logs_by_organization(client: TestClient):
    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    first_organization = create_school_by_super_admin(client, "audit-one", status="pending")
    second_organization = create_school_by_super_admin(client, "audit-two", status="pending")

    for organization in [first_organization, second_organization]:
        response = client.post(
            f"/admin/organizations/{organization['id']}/activate",
            json={"notes": "Pago externo confirmado"},
            headers=super_admin_headers,
        )
        assert response.status_code == 200, response.text

    all_logs = client.get("/admin/audit-logs", headers=super_admin_headers)
    assert all_logs.status_code == 200, all_logs.text
    assert {item["organization_id"] for item in all_logs.json()} >= {
        first_organization["id"],
        second_organization["id"],
    }
    assert all_logs.json()[0]["organization_name"] is not None
    assert all_logs.json()[0]["user_email"] == "superadmin@example.com"

    first_logs = client.get(
        "/admin/audit-logs",
        params={"organization_id": first_organization["id"], "action": "activate_organization"},
        headers=super_admin_headers,
    )
    assert first_logs.status_code == 200, first_logs.text
    assert first_logs.json()
    assert {item["organization_id"] for item in first_logs.json()} == {first_organization["id"]}
    assert {item["action"] for item in first_logs.json()} == {"activate_organization"}

    school_headers = auth_headers(client, "admin-audit-one@example.com", "SchoolAdmin123!")
    forbidden = client.get("/admin/audit-logs", headers=school_headers)
    assert forbidden.status_code == 403


def test_courses_guardians_students_and_guardian_assignment(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, guardian, student = create_school_basics(client, headers)

    assert client.get("/courses", headers=headers).json()[0]["id"] == course["id"]
    assert client.get("/guardians", headers=headers).json()[0]["phone"] == "8095551234"
    assert client.get("/students", headers=headers).json()[0]["id"] == student["id"]

    updated_course = client.put(
        f"/courses/{course['id']}",
        json={"name": "Primero", "section": "B", "academic_year": "2026-2027"},
        headers=headers,
    )
    assert updated_course.status_code == 200, updated_course.text
    assert updated_course.json()["section"] == "B"

    updated_guardian = client.put(
        f"/guardians/{guardian['id']}",
        json={"phone": "809 555 9999", "relationship": "Tutor legal"},
        headers=headers,
    )
    assert updated_guardian.status_code == 200, updated_guardian.text
    assert updated_guardian.json()["phone"] == "8095559999"

    updated_student = client.put(
        f"/students/{student['id']}",
        json={"student_code": "ST-002"},
        headers=headers,
    )
    assert updated_student.status_code == 200, updated_student.text
    assert updated_student.json()["student_code"] == "ST-002"

    second_guardian = client.post(
        "/guardians",
        json={"first_name": "Carlos", "last_name": "Pérez", "phone": "809-555-5678", "relationship": "Padre"},
        headers=headers,
    ).json()
    assignment = client.post(
        f"/students/{student['id']}/guardians",
        json={"guardian_id": second_guardian["id"], "is_primary": True},
        headers=headers,
    )
    assert assignment.status_code == 201, assignment.text
    relations = client.get(f"/students/{student['id']}/guardians", headers=headers).json()
    assert len(relations) == 2
    assert len([relation for relation in relations if relation["is_primary"]]) == 1
    assert [relation for relation in relations if relation["is_primary"]][0]["guardian_id"] == second_guardian["id"]

    removed_primary_relation = client.delete(
        f"/students/{student['id']}/guardians/{second_guardian['id']}",
        headers=headers,
    )
    assert removed_primary_relation.status_code == 204, removed_primary_relation.text
    relations_after_removal = client.get(f"/students/{student['id']}/guardians", headers=headers).json()
    assert len(relations_after_removal) == 1
    assert relations_after_removal[0]["guardian_id"] == guardian["id"]
    assert relations_after_removal[0]["is_primary"] is True

    last_relation_removal = client.delete(
        f"/students/{student['id']}/guardians/{guardian['id']}",
        headers=headers,
    )
    assert last_relation_removal.status_code == 400
    assert "al menos un tutor" in last_relation_removal.json()["detail"]

    deleted_student = client.delete(f"/students/{student['id']}", headers=headers)
    assert deleted_student.status_code == 200, deleted_student.text
    assert deleted_student.json()["is_active"] is False

    reactivated_student = client.post(f"/students/{student['id']}/reactivate", headers=headers)
    assert reactivated_student.status_code == 200, reactivated_student.text
    assert reactivated_student.json()["is_active"] is True

    duplicate_reactivation = client.post(f"/students/{student['id']}/reactivate", headers=headers)
    assert duplicate_reactivation.status_code == 409
    assert "ya está activo" in duplicate_reactivation.json()["detail"]

    deleted_guardian = client.delete(f"/guardians/{second_guardian['id']}", headers=headers)
    assert deleted_guardian.status_code == 200, deleted_guardian.text
    assert deleted_guardian.json()["is_active"] is False

    deleted_course = client.delete(f"/courses/{course['id']}", headers=headers)
    assert deleted_course.status_code == 200, deleted_course.text
    assert deleted_course.json()["is_active"] is False


def test_guardian_search_matches_name_phone_and_relationship(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    create_school_basics(client, headers)
    carlos = client.post(
        "/guardians",
        json={"first_name": "Carlos", "last_name": "Pérez", "phone": "809-555-5678", "relationship": "Padre"},
        headers=headers,
    ).json()

    by_name = client.get("/guardians", params={"search": "Carlos"}, headers=headers)
    assert by_name.status_code == 200, by_name.text
    assert [guardian["id"] for guardian in by_name.json()] == [carlos["id"]]

    by_phone = client.get("/guardians", params={"search": "8095555678"}, headers=headers)
    assert by_phone.status_code == 200, by_phone.text
    assert [guardian["id"] for guardian in by_phone.json()] == [carlos["id"]]

    by_relationship = client.get("/guardians", params={"search": "Padre"}, headers=headers)
    assert by_relationship.status_code == 200, by_relationship.text
    assert [guardian["id"] for guardian in by_relationship.json()] == [carlos["id"]]


def test_school_admin_manages_staff_and_staff_permissions(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    admin_headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, _, student = create_school_basics(client, admin_headers)

    create_staff = client.post(
        "/users",
        json={
            "first_name": "Auxiliar",
            "last_name": "Asistencia",
            "email": "staff-one@example.com",
            "password": "Staff12345!",
        },
        headers=admin_headers,
    )
    assert create_staff.status_code == 201, create_staff.text
    staff_user = create_staff.json()
    assert staff_user["role"] == "staff"
    assert staff_user["organization_id"] == organization["id"]

    listed_staff = client.get("/users", params={"role": "staff", "search": "Auxiliar"}, headers=admin_headers)
    assert listed_staff.status_code == 200, listed_staff.text
    assert [user["id"] for user in listed_staff.json()] == [staff_user["id"]]

    staff_headers = auth_headers(client, "staff-one@example.com", "Staff12345!")
    assert client.get("/courses", headers=staff_headers).status_code == 200
    forbidden_course = client.post(
        "/courses",
        json={"name": "Tercero", "section": "C", "academic_year": "2026-2027"},
        headers=staff_headers,
    )
    assert forbidden_course.status_code == 403

    attendance = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": date.today().isoformat(),
            "status": "arrived",
            "arrival_time": "07:45:00",
        },
        headers=staff_headers,
    )
    assert attendance.status_code == 201, attendance.text
    assert attendance.json()["recorded_by_user_id"] == staff_user["id"]

    staff_cannot_manage_users = client.post(
        "/users",
        json={
            "first_name": "Otro",
            "last_name": "Personal",
            "email": "staff-two@example.com",
            "password": "Staff12345!",
        },
        headers=staff_headers,
    )
    assert staff_cannot_manage_users.status_code == 403

    disabled_staff = client.delete(f"/users/{staff_user['id']}", headers=admin_headers)
    assert disabled_staff.status_code == 200, disabled_staff.text
    assert disabled_staff.json()["is_active"] is False

    blocked_login = client.post("/auth/login", json={"email": "staff-one@example.com", "password": "Staff12345!"})
    assert blocked_login.status_code == 401


def test_school_admin_uploads_logo_and_rejects_invalid_file(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")

    logo_response = client.post(
        "/organization/settings/logo",
        files={"file": ("logo.png", b"\x89PNG\r\n\x1a\nlogo-bytes", "image/png")},
        headers=headers,
    )
    assert logo_response.status_code == 200, logo_response.text
    assert logo_response.json()["logo_url"].startswith("/uploads/logos/")
    assert logo_response.json()["logo_url"].endswith(".png")
    stored_logo = client.get(logo_response.json()["logo_url"])
    assert stored_logo.status_code == 200
    assert stored_logo.content.startswith(b"\x89PNG")

    invalid_response = client.post(
        "/organization/settings/logo",
        files={"file": ("logo.txt", b"not-an-image", "text/plain")},
        headers=headers,
    )
    assert invalid_response.status_code == 400
    assert "imagen" in invalid_response.json()["detail"].lower()

    spoofed_response = client.post(
        "/organization/settings/logo",
        files={"file": ("logo.png", b"not-a-real-png", "image/png")},
        headers=headers,
    )
    assert spoofed_response.status_code == 400
    assert "contenido" in spoofed_response.json()["detail"].lower()


def test_super_admin_uploads_logo_for_created_school(client: TestClient):
    organization = create_school_by_super_admin(client)
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")

    response = client.post(
        f"/admin/organizations/{organization['id']}/logo",
        files={"file": ("school.webp", b"RIFF----WEBPlogo-bytes", "image/webp")},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["id"] == organization["id"]
    assert response.json()["logo_url"].startswith("/uploads/logos/")
    assert response.json()["logo_url"].endswith(".webp")


def test_organization_footer_is_configurable_by_super_admin_and_school_admin(client: TestClient):
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    payload = {
        "organization_name": "Colegio Footer",
        "organization_email": "footer@example.com",
        "organization_phone": "(809) 555-1212",
        "admin_first_name": "Admin",
        "admin_last_name": "Footer",
        "admin_email": "admin-footer@example.com",
        "password": "SchoolAdmin123!",
        "status": "active",
        "footer_text": "Gracias por confiar en Colegio Footer.",
    }
    response = client.post("/admin/organizations", json=payload, headers=headers)
    assert response.status_code == 201, response.text
    assert response.json()["organization"]["footer_text"] == "Gracias por confiar en Colegio Footer."

    school_headers = auth_headers(client, "admin-footer@example.com", "SchoolAdmin123!")
    settings = client.get("/organization/settings", headers=school_headers)
    assert settings.status_code == 200, settings.text
    assert settings.json()["footer_text"] == "Gracias por confiar en Colegio Footer."

    updated = client.put(
        "/organization/settings",
        json={"footer_text": "Pie actualizado para familias y personal."},
        headers=school_headers,
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["footer_text"] == "Pie actualizado para familias y personal."


def test_super_admin_updates_registered_organization(client: TestClient):
    first_organization = create_school_by_super_admin(client, "editable")
    second_organization = create_school_by_super_admin(client, "existing")
    headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")

    response = client.put(
        f"/admin/organizations/{first_organization['id']}",
        json={
            "organization_name": "Colegio Actualizado",
            "organization_email": "actualizado@example.com",
            "organization_phone": "(809) 222-3344",
            "footer_text": "Footer actualizado desde superadmin.",
            "primary_color": "#111827",
            "secondary_color": "#0F766E",
            "accent_color": "#F97316",
        },
        headers=headers,
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["name"] == "Colegio Actualizado"
    assert data["email"] == "actualizado@example.com"
    assert data["phone"] == "8092223344"
    assert data["footer_text"] == "Footer actualizado desde superadmin."
    assert data["primary_color"] == "#111827"

    duplicate = client.put(
        f"/admin/organizations/{first_organization['id']}",
        json={"organization_email": second_organization["email"]},
        headers=headers,
    )
    assert duplicate.status_code == 409

    school_headers = auth_headers(client, "admin-editable@example.com", "SchoolAdmin123!")
    forbidden = client.put(
        f"/admin/organizations/{first_organization['id']}",
        json={"organization_name": "Cambio no permitido"},
        headers=school_headers,
    )
    assert forbidden.status_code == 403


def test_super_admin_resets_school_admin_password(client: TestClient):
    organization = create_school_by_super_admin(client, "password")
    activate_school(client, organization["id"])
    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    old_school_headers = auth_headers(client, "admin-password@example.com", "SchoolAdmin123!")

    forbidden = client.put(
        f"/admin/organizations/{organization['id']}/school-admin-password",
        json={"password": "NewSchoolAdmin123!"},
        headers=old_school_headers,
    )
    assert forbidden.status_code == 403

    response = client.put(
        f"/admin/organizations/{organization['id']}/school-admin-password",
        json={"password": "NewSchoolAdmin123!"},
        headers=super_admin_headers,
    )
    assert response.status_code == 200, response.text
    assert response.json()["message"] == "Contraseña del administrador del centro actualizada."
    assert response.json()["admin_user"]["email"] == "admin-password@example.com"
    assert "password_hash" not in response.json()["admin_user"]

    old_token_me = client.get("/auth/me", headers=old_school_headers)
    assert old_token_me.status_code == 401

    old_login = client.post(
        "/auth/login",
        json={"email": "admin-password@example.com", "password": "SchoolAdmin123!"},
    )
    assert old_login.status_code == 401

    new_headers = auth_headers(client, "admin-password@example.com", "NewSchoolAdmin123!")
    me = client.get("/auth/me", headers=new_headers)
    assert me.status_code == 200, me.text
    assert me.json()["role"] == "school_admin"


def test_students_export_and_import_excel(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, _, student = create_school_basics(client, headers)

    export_response = client.get("/students/export", headers=headers)
    assert export_response.status_code == 200, export_response.text
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(export_response.content))
    worksheet = workbook.active
    assert [cell.value for cell in worksheet[1]] == [
        "primer_nombre",
        "segundo_nombre",
        "primer_apellido",
        "segundo_apellido",
        "codigo",
        "curso",
        "seccion",
        "anio_academico",
        "activo",
        "tutor_principal",
        "tutor_principal_telefono",
    ]
    assert worksheet["A1"].fill.fgColor.rgb == "FF16A34A"
    assert worksheet["A1"].font.bold is True
    assert worksheet.column_dimensions["A"].width >= len("primer_nombre") + 2
    assert worksheet.column_dimensions["K"].width >= len("tutor_principal_telefono") + 2
    assert worksheet["A2"].alignment.wrap_text is True
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert ("Luis", None, "Pérez", None, "ST-001", "Primero", "A", "2026-2027", "si", "María Rodríguez", "8095551234") in rows

    import_workbook = Workbook()
    import_sheet = import_workbook.active
    import_sheet.append([
        "primer_nombre",
        "segundo_nombre",
        "primer_apellido",
        "segundo_apellido",
        "codigo",
        "curso",
        "seccion",
        "anio_academico",
        "activo",
        "tutor_principal_telefono",
    ])
    import_sheet.append(["Luis", None, "Pérez", "Actualizado", "ST-001", "Primero", "A", "2026-2027", "si", "8095551234"])
    import_sheet.append(["Ana", None, "Gómez", None, "ST-777", "Primero", "A", "2026-2027", "no", "8095551234"])
    import_stream = BytesIO()
    import_workbook.save(import_stream)
    import_stream.seek(0)

    import_response = client.post(
        "/students/import",
        files={
            "file": (
                "estudiantes.xlsx",
                import_stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert import_response.status_code == 200, import_response.text
    assert import_response.json() == {"created": 1, "updated": 1, "errors": []}

    students_response = client.get("/students", headers=headers)
    assert students_response.status_code == 200, students_response.text
    students_by_code = {item["student_code"]: item for item in students_response.json()}
    assert students_by_code["ST-001"]["id"] == student["id"]
    assert students_by_code["ST-001"]["full_name"] == "Luis Pérez Actualizado"
    assert students_by_code["ST-777"]["course_id"] == course["id"]
    assert students_by_code["ST-777"]["is_active"] is False

    csv_export = client.get("/students/export", params={"file_format": "csv"}, headers=headers)
    assert csv_export.status_code == 200, csv_export.text
    assert csv_export.headers["content-type"].startswith("text/csv")
    assert csv_export.content.startswith(b"\xef\xbb\xbf")
    csv_rows = list(csv.DictReader(StringIO(csv_export.content.decode("utf-8-sig"))))
    assert csv_rows[0]["primer_nombre"] == "Ana"
    assert csv_rows[0]["primer_apellido"] == "Gómez"
    assert csv_rows[0]["tutor_principal_telefono"] == "8095551234"

    csv_payload = StringIO()
    writer = csv.DictWriter(csv_payload, fieldnames=[
        "primer_nombre",
        "segundo_nombre",
        "primer_apellido",
        "segundo_apellido",
        "codigo",
        "curso",
        "seccion",
        "anio_academico",
        "activo",
        "tutor_principal_telefono",
    ])
    writer.writeheader()
    writer.writerow({
        "primer_nombre": "Luis",
        "segundo_nombre": "",
        "primer_apellido": "Pérez",
        "segundo_apellido": "CSV",
        "codigo": "ST-001",
        "curso": "Primero",
        "seccion": "A",
        "anio_academico": "2026-2027",
        "activo": "si",
        "tutor_principal_telefono": "8095551234",
    })
    writer.writerow({
        "primer_nombre": "Carlos",
        "segundo_nombre": "",
        "primer_apellido": "CSV",
        "segundo_apellido": "",
        "codigo": "ST-888",
        "curso": "Primero",
        "seccion": "A",
        "anio_academico": "2026-2027",
        "activo": "si",
        "tutor_principal_telefono": "8095551234",
    })
    csv_import = client.post(
        "/students/import",
        files={"file": ("estudiantes.csv", csv_payload.getvalue().encode("utf-8-sig"), "text/csv")},
        headers=headers,
    )
    assert csv_import.status_code == 200, csv_import.text
    assert csv_import.json() == {"created": 1, "updated": 1, "errors": []}

    students_after_csv = client.get("/students", headers=headers).json()
    csv_students_by_code = {item["student_code"]: item for item in students_after_csv}
    assert csv_students_by_code["ST-001"]["full_name"] == "Luis Pérez CSV"
    assert csv_students_by_code["ST-888"]["full_name"] == "Carlos CSV"


def test_attendance_is_unique_per_student_day_and_generates_whatsapp_link(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    _, _, student = create_school_basics(client, headers)

    payload = {
        "student_id": student["id"],
        "attendance_date": date.today().isoformat(),
        "status": "arrived",
        "arrival_time": "07:45:00",
    }
    first = client.post("/attendance", json=payload, headers=headers)
    assert first.status_code == 201, first.text
    duplicate = client.post("/attendance", json=payload, headers=headers)
    assert duplicate.status_code == 409

    link = client.post(f"/attendance/{first.json()['id']}/whatsapp-link", headers=headers)
    assert link.status_code == 200, link.text
    assert link.json()["phone"] == "8095551234"
    assert link.json()["url"].startswith("https://wa.me/8095551234?text=")


def test_attendance_allows_second_record_only_for_early_pickup(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    _, _, student = create_school_basics(client, headers)
    attendance_date = date.today().isoformat()

    first = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": attendance_date,
            "status": "arrived",
            "arrival_time": "07:45:00",
        },
        headers=headers,
    )
    assert first.status_code == 201, first.text

    early_pickup = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": attendance_date,
            "status": "early_pickup",
            "departure_time": "11:30:00",
            "notes": "Retirado por su tutor principal.",
        },
        headers=headers,
    )
    assert early_pickup.status_code == 201, early_pickup.text

    duplicate_early_pickup = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": attendance_date,
            "status": "early_pickup",
            "departure_time": "12:15:00",
        },
        headers=headers,
    )
    assert duplicate_early_pickup.status_code == 409

    third_regular_record = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": attendance_date,
            "status": "late",
            "arrival_time": "08:20:00",
        },
        headers=headers,
    )
    assert third_regular_record.status_code == 409

    records = client.get("/attendance", params={"student_id": student["id"], "attendance_date": attendance_date}, headers=headers)
    assert records.status_code == 200, records.text
    assert {record["status"] for record in records.json()} == {"arrived", "early_pickup"}


def test_attendance_accepts_excused_status_and_default_template(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    _, _, student = create_school_basics(client, headers)

    templates = client.get("/whatsapp/templates", headers=headers)
    assert templates.status_code == 200, templates.text
    assert "excused" in {template["status"] for template in templates.json()}

    attendance = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": date.today().isoformat(),
            "status": "excused",
            "notes": "Excusa medica entregada.",
        },
        headers=headers,
    )
    assert attendance.status_code == 201, attendance.text
    assert attendance.json()["status"] == "excused"

    filtered = client.get("/attendance", params={"status_filter": "excused"}, headers=headers)
    assert filtered.status_code == 200, filtered.text
    assert [record["id"] for record in filtered.json()] == [attendance.json()["id"]]

    link = client.post(f"/attendance/{attendance.json()['id']}/whatsapp-link", headers=headers)
    assert link.status_code == 200, link.text
    assert "excusado" in link.json()["message"].lower()


def test_attendance_reports_apply_excuse_conversion_and_risk_colors(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, _, student = create_school_basics(client, headers)

    attendance_payloads = [
        {"attendance_date": "2026-05-01", "status": "absent"},
        {"attendance_date": "2026-05-02", "status": "absent"},
        {"attendance_date": "2026-05-03", "status": "absent"},
        {"attendance_date": "2026-05-04", "status": "excused"},
        {"attendance_date": "2026-05-05", "status": "excused"},
        {"attendance_date": "2026-05-06", "status": "excused"},
        {"attendance_date": "2026-05-07", "status": "late", "arrival_time": "08:15:00"},
    ]
    for payload in attendance_payloads:
        response = client.post(
            "/attendance",
            json={"student_id": student["id"], **payload},
            headers=headers,
        )
        assert response.status_code == 201, response.text

    student_report = client.get(
        "/reports/attendance/students",
        params={"start_date": "2026-05-01", "end_date": "2026-05-31"},
        headers=headers,
    )
    assert student_report.status_code == 200, student_report.text
    report_row = student_report.json()[0]
    assert report_row["student_id"] == student["id"]
    assert report_row["absent_count"] == 3
    assert report_row["excused_count"] == 3
    assert report_row["excused_absence_equivalent"] == 1
    assert report_row["equivalent_absences"] == 4
    assert report_row["risk_level"] == "warning"
    assert report_row["risk_color"] == "amber"
    assert report_row["records"][0]["attendance_date"] == "2026-05-01"
    assert report_row["records"][0]["status"] == "absent"
    assert report_row["records"][0]["display_time"] is None
    assert report_row["records"][-1]["attendance_date"] == "2026-05-07"
    assert report_row["records"][-1]["status"] == "late"
    assert report_row["records"][-1]["display_time"] == "08:15:00"

    course_report = client.get(
        "/reports/attendance/courses",
        params={"start_date": "2026-05-01", "end_date": "2026-05-31"},
        headers=headers,
    )
    assert course_report.status_code == 200, course_report.text
    course_row = course_report.json()[0]
    assert course_row["course_id"] == course["id"]
    assert course_row["student_count"] == 1
    assert course_row["equivalent_absences"] == 4
    assert course_row["warning_students"] == 1
    assert course_row["danger_students"] == 0
    assert course_row["risk_level"] == "warning"
    assert course_row["risk_color"] == "amber"
    assert len(course_row["records"]) == 7
    assert course_row["records"][0]["student_name"] == student["full_name"]
    assert course_row["records"][0]["attendance_date"] == "2026-05-01"
    assert course_row["records"][0]["status"] == "absent"


def test_attendance_reports_export_institutional_excel(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, _, student = create_school_basics(client, headers)

    for payload in [
        {"attendance_date": "2026-05-01", "status": "absent"},
        {"attendance_date": "2026-05-02", "status": "late", "arrival_time": "08:10:00"},
        {"attendance_date": "2026-05-03", "status": "excused"},
        {"attendance_date": "2026-05-04", "status": "excused"},
        {"attendance_date": "2026-05-05", "status": "excused"},
    ]:
        response = client.post("/attendance", json={"student_id": student["id"], **payload}, headers=headers)
        assert response.status_code == 201, response.text

    params = {"start_date": "2026-05-01", "end_date": "2026-05-31"}
    students_export = client.get("/reports/attendance/students/export", params=params, headers=headers)
    assert students_export.status_code == 200, students_export.text
    assert students_export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "reporte-asistencia-estudiantes.xlsx" in students_export.headers["content-disposition"]

    students_workbook = load_workbook(BytesIO(students_export.content))
    assert students_workbook.sheetnames == ["Resumen por estudiante", "Detalle"]
    students_summary = students_workbook["Resumen por estudiante"]
    assert students_summary["A1"].value == "Reporte institucional de asistencia"
    assert students_summary["A2"].value == organization["name"]
    assert students_summary["A6"].value == "Estudiante"
    assert students_summary["A6"].fill.fgColor.rgb == "FF2563EB"
    assert students_summary["A7"].value == student["full_name"]
    assert students_summary["D7"].value == 2
    assert students_summary["G7"].value == 1
    assert students_summary["I7"].value == "Bajo"

    students_detail = students_workbook["Detalle"]
    detail_rows = list(students_detail.iter_rows(min_row=7, values_only=True))
    assert ("2026-05-02", student["full_name"], "Primero A 2026-2027", "Tarde", "08:10", None) in detail_rows

    courses_export = client.get("/reports/attendance/courses/export", params=params, headers=headers)
    assert courses_export.status_code == 200, courses_export.text
    assert "reporte-asistencia-cursos.xlsx" in courses_export.headers["content-disposition"]

    courses_workbook = load_workbook(BytesIO(courses_export.content))
    assert courses_workbook.sheetnames == ["Resumen por curso", "Detalle"]
    courses_summary = courses_workbook["Resumen por curso"]
    assert courses_summary["A1"].value == "Reporte institucional de asistencia"
    assert courses_summary["A7"].value == "Primero A 2026-2027"
    assert courses_summary["B7"].value == 1
    assert courses_summary["C7"].value == 2
    assert courses_summary["H7"].value == 0
    assert courses_summary["I7"].value == 0

    students_pdf = client.get(
        "/reports/attendance/students/export",
        params={**params, "file_format": "pdf"},
        headers=headers,
    )
    assert students_pdf.status_code == 200, students_pdf.text
    assert students_pdf.headers["content-type"].startswith("application/pdf")
    assert "reporte-asistencia-estudiantes.pdf" in students_pdf.headers["content-disposition"]
    assert students_pdf.content.startswith(b"%PDF")
    assert len(students_pdf.content) > 1000

    courses_pdf = client.get(
        "/reports/attendance/courses/export",
        params={**params, "file_format": "pdf"},
        headers=headers,
    )
    assert courses_pdf.status_code == 200, courses_pdf.text
    assert courses_pdf.headers["content-type"].startswith("application/pdf")
    assert "reporte-asistencia-cursos.pdf" in courses_pdf.headers["content-disposition"]
    assert courses_pdf.content.startswith(b"%PDF")
    assert len(courses_pdf.content) > 1000


def test_parent_logs_in_with_phone_and_sees_only_their_students_attendance(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    school_headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    course, guardian, student = create_school_basics(client, school_headers)
    other_guardian = client.post(
        "/guardians",
        json={"first_name": "Tutor", "last_name": "Externo", "phone": "809-555-0000", "relationship": "Tutor legal"},
        headers=school_headers,
    ).json()
    other_student = client.post(
        "/students",
        json={
            "first_name": "Estudiante",
            "last_name": "No Asociado",
            "student_code": "ST-999",
            "course_id": course["id"],
            "guardian_ids": [other_guardian["id"]],
            "primary_guardian_id": other_guardian["id"],
        },
        headers=school_headers,
    ).json()

    first_attendance = client.post(
        "/attendance",
        json={
            "student_id": student["id"],
            "attendance_date": "2026-05-20",
            "status": "late",
            "arrival_time": "08:05:00",
        },
        headers=school_headers,
    )
    assert first_attendance.status_code == 201, first_attendance.text
    other_attendance = client.post(
        "/attendance",
        json={
            "student_id": other_student["id"],
            "attendance_date": "2026-05-20",
            "status": "absent",
        },
        headers=school_headers,
    )
    assert other_attendance.status_code == 201, other_attendance.text

    login = client.post("/parents/login", json={"phone": "(809) 555-1234"})
    assert login.status_code == 200, login.text
    assert "reysoft_asistencia_parent_token" in login.headers["set-cookie"]
    assert "reysoft_asistencia_csrf_token" in login.headers["set-cookie"]
    assert "HttpOnly" in login.headers["set-cookie"]
    assert login.json()["guardian"]["id"] == guardian["id"]
    assert login.json()["guardian"]["phone"] == "8095551234"
    parent_headers = {"Authorization": f"Bearer {login.json()['access_token']}"}
    parent_csrf_token = client.cookies.get("reysoft_asistencia_csrf_token")
    assert parent_csrf_token

    me = client.get("/parents/me")
    assert me.status_code == 200, me.text
    assert me.json()["full_name"] == guardian["full_name"]

    students = client.get("/parents/students", headers=parent_headers)
    assert students.status_code == 200, students.text
    assert [item["id"] for item in students.json()] == [student["id"]]
    assert students.json()[0]["course_name"] == course["name"]

    attendance = client.get("/parents/attendance", headers=parent_headers)
    assert attendance.status_code == 200, attendance.text
    assert [item["student_id"] for item in attendance.json()] == [student["id"]]
    assert attendance.json()[0]["status"] == "late"
    assert attendance.json()[0]["display_time"] == "08:05:00"

    blocked = client.get("/parents/attendance", params={"student_id": other_student["id"]}, headers=parent_headers)
    assert blocked.status_code == 404

    wrong_phone = client.post("/parents/login", json={"phone": "809-000-0000"})
    assert wrong_phone.status_code == 401

    logout = client.post("/parents/logout", headers={"X-CSRF-Token": parent_csrf_token})
    assert logout.status_code == 200, logout.text
    assert "reysoft_asistencia_parent_token" in logout.headers["set-cookie"]
    assert "reysoft_asistencia_csrf_token" in logout.headers["set-cookie"]
    assert "Max-Age=0" in logout.headers["set-cookie"]
    blocked_after_logout = client.get("/parents/me")
    assert blocked_after_logout.status_code == 401


def test_expired_activation_auto_suspends_school_and_blocks_parent_login(client: TestClient):
    organization = create_school_by_super_admin(client)
    activate_school(client, organization["id"])
    school_headers = auth_headers(client, "admin-one@example.com", "SchoolAdmin123!")
    create_school_basics(client, school_headers)

    super_admin_headers = auth_headers(client, "superadmin@example.com", "SuperAdmin123!")
    expired_date = (date.today() - timedelta(days=1)).isoformat()
    activation = client.post(
        f"/admin/organizations/{organization['id']}/activate",
        json={"expiration_date": expired_date, "notes": "Renovacion vencida."},
        headers=super_admin_headers,
    )
    assert activation.status_code == 200, activation.text

    blocked_parent_login = client.post("/parents/login", json={"phone": "(809) 555-1234"})
    assert blocked_parent_login.status_code == 401
    assert "expir" in blocked_parent_login.json()["detail"].lower()

    organization_detail = client.get(f"/admin/organizations/{organization['id']}", headers=super_admin_headers)
    assert organization_detail.status_code == 200, organization_detail.text
    assert organization_detail.json()["status"] == "suspended"


def test_school_cannot_access_other_organization_data(client: TestClient):
    first_org = create_school_by_super_admin(client, "first")
    second_org = create_school_by_super_admin(client, "second")
    activate_school(client, first_org["id"])
    activate_school(client, second_org["id"])
    first_headers = auth_headers(client, "admin-first@example.com", "SchoolAdmin123!")
    second_headers = auth_headers(client, "admin-second@example.com", "SchoolAdmin123!")

    course = client.post(
        "/courses",
        json={"name": "Segundo", "section": "B", "academic_year": "2026-2027"},
        headers=first_headers,
    ).json()
    response = client.get(f"/courses/{course['id']}", headers=second_headers)
    assert response.status_code == 404
